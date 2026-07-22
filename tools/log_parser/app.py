"""Flask app: visualize & export ECM worker logs."""

import csv
import io

from flask import Flask, jsonify, render_template, request, send_file

from parser import COLUMNS, parse_log

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 128 * 1024 * 1024  # 128 MB uploads

EXPORT_KEYS = [k for k, _ in COLUMNS]
EXPORT_HEADERS = [h for _, h in COLUMNS]


@app.route("/")
def index():
    return render_template("index.html", columns=COLUMNS)


@app.route("/parse", methods=["POST"])
def parse():
    file = request.files.get("logfile")
    if file is None or file.filename == "":
        return jsonify({"error": "No file uploaded"}), 400
    text = file.read().decode("utf-8", errors="replace")
    rows = parse_log(text)
    return jsonify({"rows": rows, "columns": COLUMNS})


def _rows_matrix(rows):
    yield EXPORT_HEADERS
    for r in rows:
        yield [r.get(k, "") if r.get(k) is not None else "" for k in EXPORT_KEYS]


@app.route("/export", methods=["POST"])
def export():
    payload = request.get_json(force=True)
    rows = payload.get("rows", [])
    fmt = payload.get("format", "csv")

    if fmt == "xlsx":
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = "ECM runs"
        header_fill = PatternFill("solid", fgColor="2F5496")
        header_font = Font(color="FFFFFF", bold=True)
        for row_idx, values in enumerate(_rows_matrix(rows), start=1):
            for col_idx, val in enumerate(values, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                if row_idx == 1:
                    cell.fill = header_fill
                    cell.font = header_font
        ws.freeze_panes = "A2"
        for col_idx, header in enumerate(EXPORT_HEADERS, start=1):
            letter = ws.cell(row=1, column=col_idx).column_letter
            ws.column_dimensions[letter].width = max(12, len(header) + 2)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(
            buf,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name="ecm_runs.xlsx",
        )

    # CSV (utf-8-sig so Excel opens it with correct encoding)
    sio = io.StringIO()
    writer = csv.writer(sio)
    for values in _rows_matrix(rows):
        writer.writerow(values)
    buf = io.BytesIO(sio.getvalue().encode("utf-8-sig"))
    return send_file(
        buf,
        mimetype="text/csv",
        as_attachment=True,
        download_name="ecm_runs.csv",
    )


if __name__ == "__main__":
    import os

    port = int(os.environ.get("PORT", "9500"))
    app.run(host="::", port=port, debug=True)
